"""
批量获取搜狗联盟前天的点击量， 已经收入

"""
from bs4 import BeautifulSoup
from hashlib import md5
from tkinter import ttk
from tkinter.filedialog import askopenfilename
from threading import Thread
import requests
import json
import os
import tkinter
import tkinter.messagebox
import openpyxl
import time

BASE_DIRS = os.path.dirname(os.path.abspath(__file__))
LOG_FILE = os.path.join(BASE_DIRS, 'log.txt')
RESULT_EXCEL = os.path.join(BASE_DIRS, 'result.xlsx')


# 若快
class RClient(object):
    """
    若快识别验证码
    """

    def __init__(self, username, password):
        self.username = username
        self.password = md5(password.encode('utf-8')).hexdigest()
        self.soft_id = '112927'
        self.soft_key = 'b17a979764df4d46a8150a2e9c18a5d6'
        self.base_params = {
            'username': self.username,
            'password': self.password,
            'softid': self.soft_id,
            'softkey': self.soft_key,
        }
        self.headers = {
            'Connection': 'Keep-Alive',
            'Expect': '100-continue',
            'User-Agent': 'ben',
        }

    def rk_create(self, im, im_type=1050, timeout=60):
        """
        识别图片
        im: 图片字节
        im_type: 题目类型
        """
        params = {
            'typeid': im_type,
            'timeout': timeout,
        }
        params.update(self.base_params)
        files = {'image': ('a.jpg', im)}
        while True:
            r = requests.post('http://api.ruokuai.com/create.json', data=params, files=files, headers=self.headers)
            if r.status_code == 200:
                print(r.text)
                if r.text.find('Error') == -1:
                    return json.loads(r.text)

    def query_user(self):
        # http://api.ruokuai.com/info.json
        data = {
            'username': self.username,
            'password': self.password
        }
        r = requests.post(url='http://api.ruokuai.com/info.json', data=data, headers=self.headers)
        print(r.text)
        if r.text.find('Error') == -1:
            return True
        else:
            return False


class SouGouGui(object):
    def __init__(self):
        self.rk = None  # RClient 实例
        self.uset_list = []
        self.data_list = []  # 最终保存到excel中
        self.task = None

        self.init_gui()

    def init_gui(self):
        self.root = tkinter.Tk()
        self.root_width = 600  # 窗口款多
        self.root_height = 320  # 窗口高度
        self.init_root_wind()  # 初始化主窗口信息

        # _________若快账号密码， 登陆，导入文件，开始处理
        self.header_frame = tkinter.Frame(self.root)
        self.header_frame.pack(fill=tkinter.X, padx=12, pady=12)
        # 若快账号
        tkinter.Label(self.header_frame, text='若快账号:').pack(side=tkinter.LEFT)
        self.rk_username = tkinter.StringVar()
        tkinter.Entry(self.header_frame, textvariable=self.rk_username, width=15).pack(side=tkinter.LEFT)
        self.rk_username.set('632673016')
        # 若快密码
        tkinter.Label(self.header_frame, text='若快密码:').pack(side=tkinter.LEFT)
        self.rk_password = tkinter.StringVar()
        tkinter.Entry(self.header_frame, textvariable=self.rk_password, width=15).pack(side=tkinter.LEFT)
        self.rk_password.set('knp158169')
        # 登陆若快
        self.login_rk_btn = tkinter.Button(self.header_frame, text='登陆若快', command=self.login_rk_func)
        self.login_rk_btn.pack(side=tkinter.LEFT, padx=15)
        # 导入文件
        self.import_file = tkinter.Button(self.header_frame, text='导入用户', command=self.import_file_func)
        self.import_file.pack(side=tkinter.LEFT)
        # 开始处理
        self.start_btn = tkinter.Button(self.header_frame, text='开始抓取', command=self.start_handel)
        self.start_btn.pack(side=tkinter.LEFT, padx=15)

        # __________________底部表格
        self.result_Frame = tkinter.Frame(self.root, height=2)
        self.result_Frame.pack(fill=tkinter.X, pady=20)

        # 定义中心列表区域
        self.bj_info_table = ttk.Treeview(self.result_Frame, show="headings", columns=('用户名', '点击量', '收入'))
        self.vbar = ttk.Scrollbar(self.result_Frame, orient=tkinter.VERTICAL, command=self.bj_info_table.yview)
        self.bj_info_table.configure(yscrollcommand=self.vbar.set)
        self.bj_info_table.column("用户名", anchor="center", width=180)
        self.bj_info_table.column("点击量", anchor="center", width=200)
        self.bj_info_table.column("收入", anchor="center", width=200)
        self.bj_info_table.heading("用户名", text="用户名")
        self.bj_info_table.heading("点击量", text="点击量")
        self.bj_info_table.heading("收入", text="收入")
        self.bj_info_table.grid(row=0, column=0, sticky=tkinter.NSEW)
        self.vbar.grid(row=0, column=1, sticky=tkinter.NS)

        self.root.mainloop()

    # 初始化主窗口
    def init_root_wind(self):
        self.root.title('搜狗')
        self.center_window()  # 设置窗口剧中
        self.root.resizable(width=False, height=False)  # 禁止窗口拉伸

    # 设置窗口剧中
    def center_window(self):
        """
        设置窗口剧中
        :param width: 窗口的宽度
        :param height: 窗口的高度
        :return: 自动设置窗口剧中
        """
        screenwidth = self.root.winfo_screenwidth()
        screenheight = self.root.winfo_screenheight()
        size = '%dx%d+%d+%d' % (
            self.root_width, self.root_height, (screenwidth - self.root_width) / 2,
            (screenheight - self.root_height) / 2)
        self.root.geometry(size)

    # 登陆按钮事件
    def login_rk_func(self):
        if not self.rk_username.get() or not self.rk_password.get():
            tkinter.messagebox.showinfo('输入有误', '用户名密码不能为空！')
            return
        rk = RClient(self.rk_username.get(), self.rk_password.get())
        if not rk.query_user():
            tkinter.messagebox.showinfo('登陆失败', '用户名密码错误！')
            return
        else:
            self.rk = rk
            tkinter.messagebox.showinfo('登陆成功', '登陆成功')

    # 导入用户
    def import_file_func(self):
        if self.rk is None:
            tkinter.messagebox.showinfo('错误', '登陆后再进行操作！')
            return
        file_dirs = askopenfilename()
        if not os.path.isfile(file_dirs):
            tkinter.messagebox.showinfo('错误', '请选择文件')
            return
        if not file_dirs.endswith('.xlsx'):
            tkinter.messagebox.showinfo('错误', '请选择xlsx格式的文件')
            return
        try:
            workbook = openpyxl.load_workbook(file_dirs)
            worksheet = workbook.active
            for row in worksheet.rows:
                user = row[0].value
                pwd = row[1].value
                self.uset_list.append([user.strip(), pwd.strip()])
                self.bj_info_table.insert('', 'end', value=(user.strip(), '还未开始', '还未开始'))
            print(self.uset_list)
        except Exception as e:
            tkinter.messagebox.showerror('失败', '导入文件失败(%s)' % e)
            self.uset_list = []

    # 开始抓取按你事件
    def start_handel(self):
        if not self.uset_list:
            tkinter.messagebox.showinfo('错误', '请先导入用户密码之后再进行处理！')
            return
        if self.task:
            if self.task.isAlive():
                tkinter.messagebox.showerror('ERRO', '当前有任务正在进行，请等待任务结束')
                return
        self.task = Thread(target=self.start_task)
        self.task.start()

    def start_task(self):
        for _ in map(self.bj_info_table.delete, self.bj_info_table.get_children("")):
            pass
        if os.path.isfile(RESULT_EXCEL):
            workbook = openpyxl.load_workbook(RESULT_EXCEL)
        else:
            workbook = openpyxl.Workbook()
        worksheet = workbook.active
        try:
            count = 0
            for i in self.uset_list:
                while True:
                    print('开始登陆:%s' % str(i))
                    if count >= 18:
                        self.bj_info_table.insert('', 'end', value=('暂停6分钟', '暂停6分钟', '暂停6分钟'))
                        time.sleep(360)
                        count = 0
                        self.bj_info_table.insert('', 'end', value=('继续执行任务', '继续执行任务', '继续执行任务'))
                    try:
                        djj, sr, d3, d4, d5, d6 = self.get_sougou_info(user=i[0], pwd=i[1], rk_obj=self.rk)
                        print([i[0], i[1], djj, sr, d3, d4])
                        worksheet.append([i[0], i[1], djj, sr, d3, d4, d5, d6])
                        workbook.save(RESULT_EXCEL)
                        self.bj_info_table.insert('', 'end', value=(i[0], djj, sr))
                        break
                    except Exception as e:
                        with open(LOG_FILE, 'a+') as f:
                            f.write('[%s]浮点数据%s\n' % (i[0], e))
                        continue
                count += 1
        except Exception as e:
            tkinter.messagebox.showerror('出现异常', '程序出现异常:%s' % e)
            return
        workbook.save(RESULT_EXCEL)
        workbook.close()
        tkinter.messagebox.showinfo('完成任务', '当前所有账号已经执行完毕， 请查看当前目录下"result.xlsx"文件！')

    def get_sougou_info(self, user, pwd, rk_obj):
        while True:
            headers = {
                'Host': 'union.sogou.com',
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.92 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
                'Referer': 'http://union.sogou.com/'
            }
            session = requests.session()
            # 获取cookies
            response = session.get(url='http://union.sogou.com/', headers=headers)
            print(response.cookies)
            # 获取验证码
            img_code_response = session.get(url='http://union.sogou.com/validateCode', headers=headers)
            print(img_code_response.cookies)
            activecode = rk_obj.rk_create(img_code_response.content).get('Result')
            print(activecode)

            # 登陆
            # http://union.sogou.com/loginauth.action
            data = {
                'loginFromPage': 'homePage',
                'systemType': '1',
                'username': user,
                'password': pwd,
                'activecode': activecode
            }
            login_response = session.post(url='http://union.sogou.com/loginauth.action', headers=headers, data=data)
            if login_response.url.endswith(user.lower()):
                with open(LOG_FILE, 'a+') as f:
                    f.write('[%s]登陆成\n' % user)
                print('登陆成功')
            else:
                soup = BeautifulSoup(login_response.text, 'lxml')
                erro_info = soup.find(name='div', id='msgTip')
                info = erro_info.text
                print(info)
                if info.find('用户名或密码不正确') != -1:
                    with open(LOG_FILE, 'a+') as f:
                        f.write('[%s]用户名或密码不正确\n' % user)
                    return '用户名或密码不正确', '用户名或密码不正确', '', '', '', ''
                if info.find('验证码无效') != -1:
                    continue

            # 转到昨天的数据页面
            yestoday_response = session.post(url='http://union.sogou.com/index.action', headers=headers,
                                             data={'searchBean.timeSegment': 'yestoday'})
            soup = BeautifulSoup(yestoday_response.text, 'lxml')
            table = soup.find(name='table', id='tableTips')
            tr = table.find_all(name='tr')[1]
            td_list = tr.find_all(name='td')

            # 转到本月的数据页面
            this_year_response = session.post(url='http://union.sogou.com/index.action', headers=headers,
                                              data={'searchBean.timeSegment': 'thismonth'})
            soup = BeautifulSoup(this_year_response.text, 'lxml')
            year_table = soup.find(name='table', id='tableTips')
            year_tr = year_table.find_all(name='tr')[1]
            year_td_list = year_tr.find_all(name='td')
            # 转到上一月
            pre_yea_response = session.post(url='http://union.sogou.com/index.action', headers=headers,
                                            data={'searchBean.timeSegment': 'lastmonth'})
            soup = BeautifulSoup(pre_yea_response.text, 'lxml')
            year_table = soup.find(name='table', id='tableTips')
            year_tr = year_table.find_all(name='tr')[1]
            pre_td_list = year_tr.find_all(name='td')

            try:
                d1, d2 = td_list[-2].text, td_list[-1].text
                d3, d4 = year_td_list[-2].text, year_td_list[-1].text
                d5, d6 = pre_td_list[-2].text, pre_td_list[-1].text
            except Exception as e:
                return '数据抓取错误', '%s' % str(td_list), '', '', '', ''

            with open(LOG_FILE, 'a+') as f:
                f.write('[%s]读取到数据[%s, %s, %s, %s]\n' % (user, d1, d2, d3, d4))

            return d1, d2, d3, d4, d5, d6





if __name__ == '__main__':
    a = SouGouGui()
    # rk = RClient(username='632673016', password='knp158169')
    # print(get_sougou_info(user='yuelai10', pwd='ma817623', rk_obj=rk))

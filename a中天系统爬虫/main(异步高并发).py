from tkinter import ttk
from tkinter.filedialog import askdirectory, askopenfilename
from tkinter import messagebox
from threading import Thread
from requests.exceptions import TooManyRedirects
from concurrent.futures import ThreadPoolExecutor, wait
import requests
import openpyxl
import tkinter
import os
import json
import time


class ZTO(object):

    def __init__(self):
        self.dh_list = []  # 保存单号的列表
        self.task = None  # 线程
        self.data_info = []  # 保存获取到的信息，方便导出
        self.data_erro = []  # 保存未成功获取的单号
        self.pool = ThreadPoolExecutor(10)
        self.init_gui()

    def init_gui(self):
        self.root = tkinter.Tk()
        self.root_width = 500  # 窗口款多
        self.root_height = 480  # 窗口高度
        self.init_root_wind()  # 初始化主窗口信息

        header_frame = tkinter.Frame(self.root)
        header_frame.pack(fill=tkinter.X, pady=20, padx=20)

        header_frame_2 = tkinter.Frame(self.root)
        header_frame_2.pack(fill=tkinter.X, pady=12, padx=20)

        header_frame_3 = tkinter.Frame(self.root)
        header_frame_3.pack(fill=tkinter.X, pady=12, padx=50)

        # zto_session
        self.zto_session = tkinter.StringVar()
        tkinter.Label(header_frame, text='zto_session:').pack(side=tkinter.LEFT)
        self.zto_session_input = tkinter.Entry(header_frame, textvariable=self.zto_session, width=40)
        self.zto_session_input.pack(padx=12)

        # session
        tkinter.Label(header_frame_2, text='     session:').pack(side=tkinter.LEFT)
        self.session = tkinter.StringVar()
        self.session_input = tkinter.Entry(header_frame_2, textvariable=self.session, width=40)
        self.session_input.pack(padx=12)

        # 导入单号按钮， 短信接口获取， 查询接口获取

        # 导入单号按钮
        self.import_excel_btn = tkinter.Button(header_frame_3, text='导入单号', command=self.import_excel)
        self.import_excel_btn.pack(side=tkinter.LEFT, padx=12)

        self.select_api = tkinter.StringVar()
        tkinter.Radiobutton(header_frame_3, variable=self.select_api, text='短信接口获取', value=0).pack(side=tkinter.LEFT)
        tkinter.Radiobutton(header_frame_3, variable=self.select_api, text='查询接口获取', value=1).pack(side=tkinter.LEFT)
        self.select_api.set(0)

        # 开始检测按钮
        self.start_btn = tkinter.Button(header_frame_3, text='开始获取', command=self.start_get)
        self.start_btn.pack(side=tkinter.LEFT, padx=12)

        # 展示信息的表格
        # 商品表格的frame
        self.goods_Frame = tkinter.Frame(self.root)
        self.goods_Frame.pack(fill=tkinter.X)

        # 定义中心列表区域
        self.info = ttk.Treeview(self.goods_Frame, show="headings", columns=('单号', '联系方式'))
        self.vbar = ttk.Scrollbar(self.goods_Frame, orient=tkinter.VERTICAL, command=self.info.yview)
        self.info.configure(yscrollcommand=self.vbar.set)
        self.info.column("单号", anchor="center", width=280)
        self.info.column("联系方式", anchor="center", width=200)
        self.info.heading("单号", text="单号")
        self.info.heading("联系方式", text="联系方式")
        self.info.grid(row=0, column=0, sticky=tkinter.NSEW)
        self.vbar.grid(row=0, column=1, sticky=tkinter.NS)

        # 导出信息按钮
        self.export_excel_btn = tkinter.Button(self.root, text='导出信息', width=40, command=self.export_info)
        self.export_excel_btn.pack(pady=10)

        self.root.mainloop()

    # 初始化主窗口
    def init_root_wind(self):
        self.root.title('中天系统v2.2')

        self.center_window()  # 设置窗口剧中
        self.root.resizable(width=False, height=False)  # 禁止窗口拉伸

    # 设置窗口剧中
    def center_window(self):
        """
        设置窗口剧中
        :param width: 窗口的宽度
        :param height: 窗口的高度
        :return: 自动设置窗口剧中
        """
        screenwidth = self.root.winfo_screenwidth()
        screenheight = self.root.winfo_screenheight()
        size = '%dx%d+%d+%d' % (
            self.root_width, self.root_height, (screenwidth - self.root_width) / 2,
            (screenheight - self.root_height) / 2)
        self.root.geometry(size)

    # 导入excel表格
    def import_excel(self):
        excel_file = askopenfilename()
        if excel_file:
            if not os.path.isfile(excel_file):
                messagebox.showinfo('导入失败', '导入失败当前文件不存在！')
                return
            if not excel_file.endswith('xlsx'):
                messagebox.showinfo('导入失败', '导入文件应该是"xlsx"文件!')
                return
            self.dh_list = []
            workbook = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active
            for row in worksheet.rows:
                value = row[0].value
                if value:
                    self.dh_list.append(value)
            # temp_list = [str(row[0].value).strip() for row in worksheet.rows if str(row[0].value)]
            if self.dh_list:
                messagebox.showinfo('导入成功', '导入成功，当前导入单号数量:  %s' % len(self.dh_list))
                print(self.dh_list)
                return
            else:
                messagebox.showinfo('导入失败', '当前文件没有数据！')
                return

    # 开始获取每个单号联系方式的按钮事件
    def start_get(self):
        # 判断单号是否导入
        if not self.dh_list:
            messagebox.showerror('Erro', '请先导入单号！')
            return
        # 判断session是否输入
        if not self.zto_session.get() or not self.session.get():
            tkinter.messagebox.showerror('ERRO', 'zto_session 或 session未输入!')
            return
        # 判断是否有任务在执行
        if self.task:
            if self.task.isAlive():
                tkinter.messagebox.showerror('ERRO', '当前有任务正在进行，请等待任务结束')
                return

        if messagebox.askyesno('开始？', '确定开始获取？'):
            self.task = Thread(target=self.inner_task)
            self.task.start()

    # 线程内的函数
    def inner_task(self):
        try:
            for _ in map(self.info.delete, self.info.get_children("")):
                pass
            # export_dh_list = [self.dh_list[i:i + 500] for i in range(0, len(self.dh_list), 500)]
            export_dh_list = self.dh_list
            self.data_info = []
            self.data_erro = []
            zto_session = self.zto_session.get().strip()
            session = self.session.get().strip()
            func = self.get_info_msg
            if int(self.select_api.get()) == 1:
                func = self.get_info

            result = [self.pool.submit(func, zto_session, session, dh) for dh in export_dh_list]
            wait(result)
            messagebox.showinfo('成功', '全部单号获取完毕，请导出查看')
        except Exception as e:
            messagebox.showerror('出现BUG', 'BUG信息:%s' % e)
            self.data_info = []
            self.data_erro = []

    # 导出信息按钮事件
    def export_info(self):
        try:
            t = time.strftime('%Y-%m-%d-%H%M%S', time.localtime())
            diretory = askdirectory()
            if diretory:
                success_file = os.path.join(diretory, '%s-success.xlsx' % t)
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                for dd in self.data_info:
                    worksheet.append(dd)
                workbook.save(success_file)

                erro_file = os.path.join(diretory, '%s-erro.xlsx' % t)
                erro_workbook = openpyxl.Workbook()
                erro_sheet = erro_workbook.active
                for dd in self.data_erro:
                    erro_sheet.append(dd)
                erro_workbook.save(erro_file)
                messagebox.showinfo('导出成功', '成功：%s\n失败：%s' % (success_file, erro_file))
        except Exception as e:
            messagebox.showerror('BUG', '导出文件出现BUG:%s' % e)

    # 程序核心, 获取指定单号的联系人
    def get_info(self, zto_session, session, dh):
        headers = {
            'X-Canvas-Fingerprint': 'a5830a699863251ddd8fb1c4af8e336v',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.92 Safari/537.36'
        }
        cookies = {
            'com.zto.sessionid': zto_session,
            'SESSION': session
        }
        try:
            response = requests.get(
                url='https://sso.zto.com/security-services/billtrack/billinfo-query-preauth?bill_id=%s&type=A014' % dh,
                headers=headers, cookies=cookies, timeout=10)
            time.sleep(0.1)
            print(response.text)
            # {"error":"no_perm_query:not_scan_node","error_description":"运单不属于贵网点","message":"该运单未在贵网点进行到件、发件、中转扫描或不是发放给贵网点使用的，不能进行「订单信息」查询操作。\n如您是客服中心或服务网点用户，请向您的主管确认您所在的网点具有此运单流经路由网点的查询权限。"}
            # {"ticket":"4He3WNFTEeiX3QBQVoRltw"}
            # {"error":"you_need_login_first","error_description":"需要登录","message":"您需要登录才可以继续此操作。"}
            data = json.loads(response.text)
            if data.get('error', None):
                info = data.get('error_description')
                self.data_erro.append([dh, info])
                self.info.insert('', 0, values=(dh, info))
                return dh, info
            headers_2 = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.92 Safari/537.36',
                '_type': 'A014',
                '_token': data.get('ticket', '')
            }
            try:
                response = requests.get(url='https://newbill.zt-express.com/order-query/get?billCode=%s' % dh,
                                        headers=headers_2, cookies=cookies, timeout=10)
                time.sleep(0.1)

                print(dh, response.text)
            except TooManyRedirects as e:
                self.data_erro.append([dh, 'SESSION失效'])
                self.info.insert('', 0, values=(dh, 'SESSION失效'))
                return dh, 'SESSION失效'
            data2 = json.loads(response.text)
            if data2['status']:
                if data2.get('result', None):
                    ii = data2['result'][0]['receiveInfo']
                    if ii == "R11:没有获取到收件人信息":
                        self.data_erro.append([dh, ii])
                        self.info.insert('', 0, values=(dh, ii))
                    else:
                        self.info.insert('', 0, values=(dh, ii))
                        self.data_info.append([dh, ii])
                    return dh, ii
                else:
                    self.data_erro.append([dh, '暂无数据'])
                    self.info.insert('', 0, values=(dh, '暂无数据'))
                    return dh, '暂无数据'
            else:
                self.data_erro.append([dh, response.text])
                self.info.insert('', 0, values=(dh, response.text))
                return dh, data2
        except Exception as e:
            info = '查询接口获取单号:%s出错:%s\n' % e
            self.data_erro.append([dh, info])
            return dh, str(e)

    # 短信接口获取联系方式
    def get_info_msg(self, zto_session, session, dh):
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.92 Safari/537.36'
            }
            data = {"billCode": dh}
            cookies = {
                'com.zto.sessionid': zto_session,
                'SESSION': session
            }
            try:
                time.sleep(0.1)
                response = requests.post(url='https://sms.zto.com/mobile/billcode', json=data, cookies=cookies, timeout=10, headers=headers)
            except TooManyRedirects as e:
                self.data_erro.append([dh, 'SESSION失效'])
                self.info.insert('', 0, values=(dh, 'SESSION失效'))
                return dh, 'SESSION失效'
            info = json.loads(response.text)
            print(dh, info)
            # {"resultData":null,"requestStatus":false,"StatusMessage":"无法正确获取手机号"}
            # {"resultData":{"RecMobile":"13916437644","RecName":""},"requestStatus":true,"StatusMessage":"成功获取"}
            if info.get('requestStatus'):
                ii = info['resultData']['RecMobile']
                self.data_info.append([dh, ii])
                self.info.insert('', 0, values=(dh, ii))
                return dh, ii
            else:
                self.data_erro.append([dh, '无法正确获取手机号'])
                self.info.insert('', 0, values=(dh, '无法正确获取手机号'))
                return dh, '无法正确获取手机号'
        except Exception as e:
            info = '短信接口获取单号:%s时候出错%s\n' % (dh, e)
            self.data_erro.append([dh, info])
            return dh, str(e)


if __name__ == '__main__':
    ZTO()
